#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import logging
import datetime
import openpyxl
from openpyxl.styles import numbers
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters
)
from dotenv import load_dotenv
from PIL import Image as PILImage
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import json
import asyncio
from aiohttp import web

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Загрузка конфигурации
load_dotenv(override=True)

# Настройка порта для Railway
PORT = int(os.getenv('PORT', 8080))

# Получение переменных окружения
TOKEN = os.getenv("TELEGRAM_TOKEN")
logger.info(f"TELEGRAM_TOKEN: {'présent' if TOKEN else 'absent'}")
if not TOKEN:
    logger.error("❌ ОШИБКА: Токен Telegram не найден")
    logger.error("Проверьте переменные окружения в Railway")
    sys.exit(1)

# Настройки Google
GOOGLE_SHEETS_SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
GOOGLE_DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive']
SPREADSHEET_ID = os.getenv("GOOGLE_SHEET_ID")
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID")
logger.info(f"GOOGLE_SHEET_ID: {'présent' if SPREADSHEET_ID else 'absent'}")
logger.info(f"GOOGLE_DRIVE_FOLDER_ID: {'présent' if GOOGLE_DRIVE_FOLDER_ID else 'absent'}")

# Получение учетных данных Google из переменной окружения
GOOGLE_CREDS_JSON = os.getenv("GOOGLE_CREDS_JSON")
logger.info(f"GOOGLE_CREDS_JSON: {'présent' if GOOGLE_CREDS_JSON else 'absent'}")
if not GOOGLE_CREDS_JSON:
    logger.error("❌ ОШИБКА: Учетные данные Google не найдены")
    logger.error("Проверьте переменные окружения в Railway")
    sys.exit(1)

try:
    creds_dict = json.loads(GOOGLE_CREDS_JSON)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=GOOGLE_SHEETS_SCOPES + GOOGLE_DRIVE_SCOPES
    )
except Exception as e:
    logger.error(f"❌ ОШИБКА при загрузке учетных данных Google: {e}")
    sys.exit(1)

if not SPREADSHEET_ID or not GOOGLE_DRIVE_FOLDER_ID:
    logger.error("❌ ОШИБКА: ID таблицы или папки Drive не указаны")
    sys.exit(1)

# Локальные настройки
PHOTOS_DIR = "photos"
EXCEL_FILE = os.getenv("EXCEL_FILE", "журнал_выпечки.xlsx")
MAX_IMAGE_WIDTH = 800
MAX_IMAGE_HEIGHT = 600

# Список наименований продукции (полные названия)
PRODUCT_NAMES = [
    "Круассан с ветчиной и сыром",
    "Круассан классический",
    "Миникруассан",
    "Круассан с шоколадной начинкой",
    "Круассан с соленой карамелью",
    "Слойка с соленой карамелью",
    "Слойка дрожжевая с начинкой «Малиновая»",
    "Слойка дрожжевая с начинкой «Маковая»",
    "Слойка с сыром и шпинатом",
    "Сосиска в тесте"
]

# Создание директории для фото
os.makedirs(PHOTOS_DIR, exist_ok=True)

# Заголовки столбцов
COLUMN_HEADERS = [
    'Мастер',
    'Дата',
    'Смена',
    'Наименование',
    'Комментарий',
    'Фото'
]

def init_google_services():
    """Инициализация сервисов Google"""
    try:
        sheets_service = build('sheets', 'v4', credentials=creds)
        drive_service = build('drive', 'v3', credentials=creds)
        return sheets_service, drive_service
    except Exception as e:
        logger.error(f"❌ ОШИБКА Google: {e}")
        return None, None

def init_excel_file():
    """Инициализация файла Excel"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Журнал выпечки"
        ws.append(COLUMN_HEADERS)
        
        # Настройка ширины столбцов
        column_widths = [20, 15, 10, 20, 30, 20]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
        wb.save(EXCEL_FILE)

def save_to_excel(data, photo_path):
    """Сохранение данных в Excel"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        row_num = ws.max_row + 1
        
        # Добавление данных
        ws.cell(row=row_num, column=1, value=data['Мастер'])
        ws.cell(row=row_num, column=2, value=data['date_obj']).number_format = 'DD.MM.YYYY'
        ws.cell(row=row_num, column=3, value=data['смена'])
        ws.cell(row=row_num, column=4, value=data['наименование'])
        ws.cell(row=row_num, column=5, value=data.get('комментарий', ''))
        
        # Добавление фото
        img = XLImage(photo_path)
        img.width = 150
        img.height = 100
        img.anchor = f'F{row_num}'
        ws.add_image(img)
        
        ws.row_dimensions[row_num].height = 80
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"❌ ОШИБКА Excel: {e}")
        return False

async def upload_to_drive(file_path, drive_service):
    """Загрузка файла на Google Drive"""
    try:
        file_metadata = {
            'name': os.path.basename(file_path),
            'parents': [GOOGLE_DRIVE_FOLDER_ID]
        }
        media = MediaFileUpload(file_path, mimetype='image/jpeg')
        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        # Открываем доступ
        drive_service.permissions().create(
            fileId=file['id'],
            body={'type': 'anyone', 'role': 'reader'}
        ).execute()
        
        return f"https://drive.google.com/uc?id={file['id']}"
    except Exception as e:
        logger.error(f"❌ ОШИБКА Drive: {e}")
        return None

async def save_to_sheets(sheets_service, data, image_url):
    """Сохранение в Google Sheets"""
    try:
        # Проверка заголовков
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID,
            range="A1:F1"
        ).execute()
        
        if 'values' not in result:
            sheets_service.spreadsheets().values().update(
                spreadsheetId=SPREADSHEET_ID,
                range="A1:F1",
                valueInputOption='USER_ENTERED',
                body={'values': [COLUMN_HEADERS]}
            ).execute()

        # Форматируем дату как строку в формате ДД.ММ.ГГГГ
        date_str = data['date_obj'].strftime("%d.%m.%Y")
        
        values = [
            [
                data['Мастер'],
                date_str,
                data['смена'],
                data['наименование'],
                data.get('комментарий', ''),
                f'=IMAGE("{image_url}")'
            ]
        ]

        # Добавление данных
        sheets_service.spreadsheets().values().append(
            spreadsheetId=SPREADSHEET_ID,
            range="A1:F1",
            valueInputOption='USER_ENTERED',
            insertDataOption='INSERT_ROWS',
            body={'values': values}
        ).execute()
        
        return True
    except Exception as e:
        logger.error(f"❌ ОШИБКА Sheets: {e}")
        return False

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка команды /start"""
    context.user_data.clear()
    context.user_data['этап'] = 1
    await update.message.reply_text(
        "👨‍🍳 Введите ФИО Мастера:",
        reply_markup=ReplyKeyboardRemove()
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка текстовых сообщений"""
    user_data = context.user_data
    text = update.message.text.strip()

    if user_data['этап'] == 1:
        user_data['Мастер'] = text
        user_data['этап'] = 2
        
        # Nouvelle interface pour la date avec boutons
        keyboard = [
            [InlineKeyboardButton("Сегодня", callback_data='date_today')],
            [InlineKeyboardButton("Другая дата", callback_data='date_custom')]
        ]
        
        await update.message.reply_text(
            "📅 Выберите дату:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    elif user_data['этап'] == 2:
        # Traitement si l'utilisateur entre une date manuellement
        if 'custom_date' in user_data and user_data['custom_date']:
            try:
                date_obj = datetime.datetime.strptime(text, "%d.%m.%Y")
                user_data['дата'] = text
                user_data['date_obj'] = date_obj
                del user_data['custom_date']
                user_data['этап'] = 3
                
                await update.message.reply_text(
                    "🌞🌜 Выберите смену:",
                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton("День", callback_data='день'),
                         InlineKeyboardButton("Ночь", callback_data='ночь')]
                    ])
                )
            except ValueError:
                await update.message.reply_text("❌ Неверный формат. Используйте ДД.ММ.ГГГГ")
                return
    
    elif user_data['этап'] == 4:
        if 'custom_name' in user_data and user_data['custom_name']:
            user_data['наименование'] = text
            del user_data['custom_name']
            user_data['этап'] = 5
            await update.message.reply_text("💬 Введите комментарий:")
    
    elif user_data['этап'] == 5:
        user_data['комментарий'] = text
        user_data['этап'] = 6
        await update.message.reply_text("📸 Отправьте фото готовой продукции:")

async def handle_date_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора даты"""
    query = update.callback_query
    await query.answer()

    user_data = context.user_data
    
    if query.data == 'date_today':
        today = datetime.datetime.now()
        user_data['дата'] = today.strftime("%d.%m.%Y")
        user_data['date_obj'] = today
        user_data['этап'] = 3
        
        await query.edit_message_text(text=f"Дата: {user_data['дата']}")
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="🌞🌜 Выберите смену:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("День", callback_data='день'),
                 InlineKeyboardButton("Ночь", callback_data='ночь')]
            ])
        )
    elif query.data == 'date_custom':
        user_data['custom_date'] = True
        await query.edit_message_text(text="✏️ Введите дату в формате ДД.ММ.ГГГГ:")
    elif user_data['этап'] == 2:
        if text.lower() == 'сегодня':
            today = datetime.datetime.now()
            user_data['дата'] = today.strftime("%d.%m.%Y")
            user_data['date_obj'] = today
        else:
            try:
                date_obj = datetime.datetime.strptime(text, "%d.%m.%Y")
                user_data['дата'] = text
                user_data['date_obj'] = date_obj
            except ValueError:
                await update.message.reply_text("❌ Неверный формат. Используйте ДД.ММ.ГГГГ")
                return

        user_data['этап'] = 3
        await update.message.reply_text(
            "🌞🌜 Выберите смену:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("День", callback_data='день'),
                 InlineKeyboardButton("Ночь", callback_data='ночь')]
            ])
        )
    
    elif user_data['этап'] == 4:
        # Если пользователь вводит собственное наименование
        if 'custom_name' in user_data and user_data['custom_name']:
            user_data['наименование'] = text
            del user_data['custom_name']
            user_data['этап'] = 5
            await update.message.reply_text("💬 Введите комментарий:")
    
    elif user_data['этап'] == 5:
        user_data['комментарий'] = text
        user_data['этап'] = 6
        await update.message.reply_text("📸 Отправьте фото готовой продукции:")

async def handle_shift(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора смены"""
    query = update.callback_query
    await query.answer()

    user_data = context.user_data
    user_data['смена'] = 'День' if query.data == 'день' else 'Ночь'
    user_data['этап'] = 4

    await query.edit_message_text(text=f"Смена: {user_data['смена']}")
    
    # Создание клавиатуры с одним продуктом на строку
    keyboard = []
    for idx, name in enumerate(PRODUCT_NAMES):
        keyboard.append([InlineKeyboardButton(name, callback_data=f"name_{idx}")])
    
    # Добавление кнопки для ввода своего наименования
    keyboard.append([InlineKeyboardButton("✏️ Ввести свое наименование", callback_data="custom_name")])
    
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="🏷 Выберите наименование продукта:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def handle_product_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора наименования"""
    query = update.callback_query
    await query.answer()

    user_data = context.user_data
    
    if query.data == "custom_name":
        user_data['custom_name'] = True
        await query.edit_message_text(text="✏️ Введите свое наименование:")
        return
    
    idx = int(query.data.split("_")[1])
    user_data['наименование'] = PRODUCT_NAMES[idx]
    user_data['этап'] = 5
    
    await query.edit_message_text(text=f"Наименование: {user_data['наименование']}")
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="💬 Введите комментарий:"
    )

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка фото"""
    user_data = context.user_data
    if user_data.get('этап') != 6:
        await update.message.reply_text("❌ Начните с команды /start")
        return

    try:
        # Скачивание фото
        photo_file = await update.message.photo[-1].get_file()
        img_data = BytesIO()
        await photo_file.download_to_memory(out=img_data)
        img_data.seek(0)

        # Обработка изображения
        with PILImage.open(img_data) as img:
            img.thumbnail((MAX_IMAGE_WIDTH, MAX_IMAGE_HEIGHT))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            photo_path = os.path.join(PHOTOS_DIR, f"{timestamp}.jpg")
            img.save(photo_path, "JPEG", quality=85)

        # Сохранение
        excel_ok = save_to_excel(user_data, photo_path)
        sheets_ok = False
        image_url = None
        
        sheets_service, drive_service = init_google_services()
        if sheets_service and drive_service:
            image_url = await upload_to_drive(photo_path, drive_service)
            if image_url:
                sheets_ok = await save_to_sheets(sheets_service, user_data, image_url)

        # Формирование ответа
        if excel_ok and sheets_ok:
            msg = (
                "✅ Данные сохранены!\n"
                f"📅 Дата: {user_data['дата']}\n"
                f"👨‍🍳 Мастер: {user_data['Мастер']}\n"
                f"🌃 Смена: {user_data['смена']}\n"
                f"🏷 Наименование: {user_data['наименование']}\n"
                f"💬 Комментарий: {user_data.get('комментарий', 'нет')}\n\n"
                "Добавить новую запись?"
            )
        elif excel_ok:
            msg = "⚠️ Данные сохранены в Excel, но ошибка Google Sheets\nДобавить запись?"
        else:
            msg = "❌ Ошибка сохранения\nПопробовать снова?"

        await update.message.reply_text(
            msg,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ Новая запись", callback_data='новая')]
            ])
        )

    except Exception as e:
        logger.error(f"❌ ОШИБКА фото: {e}")
        await update.message.reply_text(
            "❌ Ошибка. Попробуйте снова.\n"
            "Начните с /start"
        )

async def handle_new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Новая запись"""
    query = update.callback_query
    await query.answer()
    await start(query, context)

async def health_check(request):
    """Endpoint de vérification de santé pour Railway"""
    return web.Response(text="OK")

async def main():
    """Fonction principale de démarrage du bot"""
    # Vérification des variables d'environnement
    logger.info("Vérification des variables d'environnement...")
    if not TOKEN:
        logger.error("❌ TELEGRAM_TOKEN manquant")
        return
    if not GOOGLE_CREDS_JSON:
        logger.error("❌ GOOGLE_CREDS_JSON manquant")
        return
    if not SPREADSHEET_ID:
        logger.error("❌ GOOGLE_SHEET_ID manquant")
        return
    if not GOOGLE_DRIVE_FOLDER_ID:
        logger.error("❌ GOOGLE_DRIVE_FOLDER_ID manquant")
        return

    logger.info("✅ Variables d'environnement OK")
    logger.info("Initialisation de l'application...")

    # Initialisation de l'application
    application = Application.builder().token(TOKEN).build()
    
    # Ajout des gestionnaires
    logger.info("Configuration des gestionnaires...")
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("new", handle_new))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(CallbackQueryHandler(handle_date_choice, pattern="^date_"))
    application.add_handler(CallbackQueryHandler(handle_shift, pattern="^день|ночь$"))
    application.add_handler(CallbackQueryHandler(handle_product_name, pattern="^name_|custom_name$"))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    
    logger.info("✅ Configuration terminée")
    logger.info("Démarrage du bot...")

    try:
        # Initialisation de l'application
        await application.initialize()
        await application.start()
        
        # Démarrage du polling
        await application.run_polling(
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True
        )
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'exécution du bot: {str(e)}")
    finally:
        # Arrêt propre de l'application
        try:
            logger.info("Arrêt de l'application...")
            await application.stop()
            await application.shutdown()
            logger.info("✅ Application arrêtée avec succès")
        except Exception as e:
            logger.error(f"❌ Erreur lors de l'arrêt du bot: {str(e)}")

def run_bot():
    """Fonction pour exécuter le bot"""
    try:
        # Création d'une nouvelle boucle d'événements
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            # Exécution de la fonction principale
            loop.run_until_complete(main())
        except KeyboardInterrupt:
            logger.info("Arrêt manuel du bot")
        except Exception as e:
            logger.error(f"❌ Erreur lors de l'exécution du bot: {str(e)}")
            sys.exit(1)
        finally:
            # Nettoyage de la boucle d'événements
            try:
                pending = asyncio.all_tasks(loop)
                loop.run_until_complete(asyncio.gather(*pending, return_exceptions=True))
                loop.run_until_complete(loop.shutdown_asyncgens())
            finally:
                loop.close()
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'exécution du bot: {str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    logger.info("🚀 Démarrage du programme...")
    run_bot()
